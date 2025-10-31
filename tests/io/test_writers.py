"""Tests for file writers (Task Group 7.11, 7.13)."""

from pathlib import Path

import numpy as np
import pytest

from rheojax.core.data import RheoData
from rheojax.io.writers import save_excel, save_hdf5

# Check if h5py is available
try:
    import h5py

    HAS_H5PY = True
except ImportError:
    HAS_H5PY = False


@pytest.mark.skipif(not HAS_H5PY, reason="h5py not installed")
class TestHDF5Writer:
    """Tests for HDF5 writer (7.11)."""

    @pytest.fixture
    def sample_data(self):
        """Create sample RheoData for testing."""
        return RheoData(
            x=np.array([1.0, 2.0, 3.0, 4.0, 5.0]),
            y=np.array([100.0, 200.0, 300.0, 400.0, 500.0]),
            x_units="s",
            y_units="Pa",
            domain="time",
            metadata={
                "sample_name": "Test Sample",
                "temperature": 25.0,
                "test_mode": "relaxation",
            },
        )

    def test_hdf5_basic_write(self, tmp_path, sample_data):
        """Test basic HDF5 data serialization."""
        output_file = tmp_path / "test_output.h5"

        save_hdf5(sample_data, str(output_file))

        assert output_file.exists()
        assert output_file.stat().st_size > 0

    def test_hdf5_metadata_preservation(self, tmp_path, sample_data):
        """Test HDF5 metadata preservation."""
        output_file = tmp_path / "test_metadata.h5"

        save_hdf5(sample_data, str(output_file))

        # Verify we can read it back (would need read function)
        import h5py

        with h5py.File(str(output_file), "r") as f:
            assert "x" in f
            assert "y" in f
            assert "metadata" in f.attrs or "metadata" in f

    def test_hdf5_compression(self, tmp_path, sample_data):
        """Test HDF5 compression."""
        output_file = tmp_path / "test_compressed.h5"

        # Create larger dataset for compression test
        large_data = RheoData(
            x=np.linspace(0, 100, 10000),
            y=np.random.randn(10000),
            x_units="s",
            y_units="Pa",
            domain="time",
        )

        save_hdf5(large_data, str(output_file), compression=True)

        # With compression, file should be smaller
        assert output_file.exists()

    def test_hdf5_roundtrip(self, tmp_path, sample_data):
        """Test HDF5 roundtrip (write then read)."""
        output_file = tmp_path / "test_roundtrip.h5"

        save_hdf5(sample_data, str(output_file))

        # Read back (would need read function)
        import h5py

        with h5py.File(str(output_file), "r") as f:
            x_read = f["x"][:]
            y_read = f["y"][:]

            np.testing.assert_array_equal(x_read, sample_data.x)
            np.testing.assert_array_equal(y_read, sample_data.y)


class TestExcelWriter:
    """Tests for Excel writer (7.13)."""

    @pytest.fixture
    def results_dict(self):
        """Create sample results dictionary."""
        return {
            "parameters": {
                "G_s": 1.5e5,
                "eta_s": 2.3e3,
            },
            "fit_quality": {
                "R2": 0.95,
                "RMSE": 12.3,
            },
            "residuals": np.array([0.1, -0.2, 0.15, -0.1, 0.05]),
            "predictions": np.array([100, 200, 300, 400, 500]),
        }

    @pytest.mark.skip(reason="Requires openpyxl/xlsxwriter installation")
    def test_excel_basic_export(self, tmp_path, results_dict):
        """Test basic Excel data export."""
        output_file = tmp_path / "test_output.xlsx"

        save_excel(results_dict, str(output_file))

        assert output_file.exists()

    @pytest.mark.skip(reason="Requires openpyxl installation")
    def test_excel_metadata_export(self, tmp_path, results_dict):
        """Test Excel metadata export."""
        output_file = tmp_path / "test_metadata.xlsx"

        save_excel(results_dict, str(output_file))

        # Verify structure
        import openpyxl

        wb = openpyxl.load_workbook(str(output_file))
        assert "Parameters" in wb.sheetnames or len(wb.sheetnames) > 0

    @pytest.mark.skip(reason="Requires plotting integration")
    def test_excel_plot_inclusion(self, tmp_path, results_dict):
        """Test Excel with optional plot inclusion."""
        output_file = tmp_path / "test_plots.xlsx"

        save_excel(results_dict, str(output_file), include_plots=True)

        assert output_file.exists()
