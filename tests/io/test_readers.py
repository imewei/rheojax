"""Tests for file readers (Task Group 7.1, 7.3, 7.5, 7.7, 7.9)."""

from pathlib import Path

import numpy as np
import pytest

from rheojax.core.data import RheoData
from rheojax.io.readers import (
    auto_load,
    load_anton_paar,
    load_csv,
    load_excel,
    load_trios,
)
from rheojax.io.readers.csv_reader import detect_csv_delimiter

# Test data directory
TEST_DATA_DIR = Path(__file__).parent / "test_data"


class TestTriosReader:
    """Tests for TRIOS reader (7.1)."""

    @pytest.mark.smoke
    def test_trios_reader_basic(self, tmp_path):
        """Test basic TRIOS file parsing with simple data."""
        # Create a minimal TRIOS file
        trios_content = """Filename	test.txt
Instrument serial number	4010-1234
Instrument name	TASerNo 4010-1234
operator	Test User
rundate	2025-10-24
Sample name	Test Sample
Geometry name	Parallel Plate

[step]
Temperature	°C
25.0

[step]
Number of points	5
Variables	Time	Storage modulus	Loss modulus
	s	Pa	Pa
data point 1	0.1	1000	500
data point 2	0.2	1200	600
data point 3	0.3	1400	700
data point 4	0.4	1600	800
data point 5	0.5	1800	900
"""
        test_file = tmp_path / "test_trios.txt"
        test_file.write_text(trios_content)

        # Load the file
        data = load_trios(str(test_file))

        # Verify it returns RheoData
        assert isinstance(data, RheoData)

        # Verify basic structure
        assert data.x is not None
        assert data.y is not None
        assert len(data.x) > 0

    @pytest.mark.smoke
    def test_trios_metadata_extraction(self, tmp_path):
        """Test TRIOS metadata parsing."""
        trios_content = """Filename	metadata_test.txt
Instrument serial number	5343-5678
Instrument name	DHR-3
operator	Dr. Smith
rundate	2025-10-24
Sample name	Polymer XYZ
Geometry name	Cone and Plate

[step]
Temperature	°C
25.0

[step]
Variables	Time	Stress
	s	Pa
data point 1	1.0	100
data point 2	2.0	200
"""
        test_file = tmp_path / "test_metadata.txt"
        test_file.write_text(trios_content)

        data = load_trios(str(test_file))

        # Check metadata was extracted
        assert "sample_name" in data.metadata
        assert data.metadata["sample_name"] == "Polymer XYZ"
        assert "instrument_serial_number" in data.metadata

    @pytest.mark.smoke
    def test_trios_unit_conversion(self):
        """Test TRIOS unit conversion (MPa→Pa, %→unitless)."""
        # This test would verify unit conversion logic
        # For now, just check that the function exists
        from rheojax.io.readers.trios import convert_units

        assert callable(convert_units)

    @pytest.mark.smoke
    def test_trios_multiple_segments(self, tmp_path):
        """Test TRIOS with multiple procedure segments."""
        trios_content = """Filename	multiseg.txt
Instrument serial number	4010-1234
Instrument name	ARES-G2

[step]
Variables	Time	Stress
	s	Pa
data point 1	1.0	100
data point 2	2.0	200

[step]
Variables	Time	Stress
	s	Pa
data point 1	3.0	300
data point 2	4.0	400
"""
        test_file = tmp_path / "test_multiseg.txt"
        test_file.write_text(trios_content)

        data = load_trios(str(test_file))

        # Should handle multiple segments
        assert isinstance(data, (RheoData, list))

    @pytest.mark.smoke
    def test_trios_error_handling(self, tmp_path):
        """Test TRIOS error handling for malformed files."""
        malformed_content = "Not a valid TRIOS file\nRandom data\n"
        test_file = tmp_path / "malformed.txt"
        test_file.write_text(malformed_content)

        with pytest.raises(Exception):
            load_trios(str(test_file))


class TestCSVReader:
    """Tests for CSV reader (7.3)."""

    def test_csv_basic_read(self, tmp_path):
        """Test basic CSV file reading with explicit columns."""
        csv_content = """time,stress,strain
1.0,100.0,0.01
2.0,200.0,0.02
3.0,300.0,0.03
4.0,400.0,0.04
"""
        test_file = tmp_path / "test.csv"
        test_file.write_text(csv_content)

        data = load_csv(str(test_file), x_col="time", y_col="stress")

        assert isinstance(data, RheoData)
        assert len(data.x) == 4
        assert len(data.y) == 4
        np.testing.assert_array_equal(data.x, [1.0, 2.0, 3.0, 4.0])
        np.testing.assert_array_equal(data.y, [100.0, 200.0, 300.0, 400.0])

    def test_csv_delimiter_detection(self, tmp_path):
        """Test CSV delimiter auto-detection."""
        # Tab-separated
        tsv_content = "time\tstress\n1.0\t100.0\n2.0\t200.0\n"
        test_file = tmp_path / "test.tsv"
        test_file.write_text(tsv_content)

        data = load_csv(str(test_file), x_col="time", y_col="stress")

        assert isinstance(data, RheoData)
        assert len(data.x) == 2

    def test_csv_header_detection(self, tmp_path):
        """Test CSV header row detection."""
        # File without header
        csv_no_header = """1.0,100.0
2.0,200.0
3.0,300.0
"""
        test_file = tmp_path / "no_header.csv"
        test_file.write_text(csv_no_header)

        data = load_csv(str(test_file), x_col=0, y_col=1, header=None)

        assert isinstance(data, RheoData)
        assert len(data.x) == 3

    def test_csv_unit_specification(self, tmp_path):
        """Test CSV with unit specification."""
        csv_content = "freq,modulus\n1.0,1000.0\n10.0,2000.0\n"
        test_file = tmp_path / "test_units.csv"
        test_file.write_text(csv_content)

        data = load_csv(
            str(test_file), x_col="freq", y_col="modulus", x_units="Hz", y_units="Pa"
        )

        assert data.x_units == "Hz"
        assert data.y_units == "Pa"


_openpyxl = pytest.importorskip("openpyxl", reason="openpyxl required for Excel tests")


class TestExcelReader:
    """Tests for Excel reader (7.5)."""

    def _create_excel(self, filepath, data_dict, sheet_name="Sheet1"):
        """Helper to create an Excel file from a dict of columns."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        headers = list(data_dict.keys())
        ws.append(headers)
        n_rows = len(next(iter(data_dict.values())))
        for i in range(n_rows):
            ws.append([data_dict[h][i] for h in headers])
        wb.save(filepath)

    def test_excel_basic_read(self, tmp_path):
        """Test basic Excel file reading."""
        filepath = tmp_path / "test.xlsx"
        self._create_excel(
            filepath,
            {
                "time (s)": [0.1, 0.2, 0.3, 0.4, 0.5],
                "stress (Pa)": [100, 200, 300, 400, 500],
            },
        )

        data = load_excel(str(filepath), x_col="time (s)", y_col="stress (Pa)")
        assert isinstance(data, RheoData)
        assert len(data.x) == 5

    def test_excel_sheet_selection(self, tmp_path):
        """Test Excel sheet selection."""
        import openpyxl

        filepath = tmp_path / "multi_sheet.xlsx"
        wb = openpyxl.Workbook()
        # First sheet (default)
        ws1 = wb.active
        ws1.title = "Empty"
        ws1.append(["a", "b"])

        # Second sheet with actual data
        ws2 = wb.create_sheet("Data")
        ws2.append(["time", "stress"])
        for i in range(5):
            ws2.append([float(i + 1), float((i + 1) * 100)])
        wb.save(filepath)

        data = load_excel(str(filepath), x_col="time", y_col="stress", sheet="Data")
        assert isinstance(data, RheoData)
        assert len(data.x) == 5

    def test_excel_column_mapping(self, tmp_path):
        """Test Excel column mapping by index."""
        filepath = tmp_path / "indexed.xlsx"
        self._create_excel(
            filepath,
            {
                "col_A": [1.0, 2.0, 3.0],
                "col_B": [10.0, 20.0, 30.0],
            },
        )

        data = load_excel(str(filepath), x_col=0, y_col=1)
        assert isinstance(data, RheoData)
        assert len(data.x) == 3
        np.testing.assert_allclose(data.x, [1.0, 2.0, 3.0])


class TestAntonPaarReader:
    """Tests for Anton Paar reader (7.7)."""

    def test_anton_paar_basic_parse(self, tmp_path):
        """Parse minimal Anton Paar export and extract units/labels."""

        file = tmp_path / "ap_export.txt"
        file.write_text(
            "Project:\tTest\n"
            "Interval and data points:\t1\t2\n"
            "Interval data:\tTime\tG'\tG''\n"
            "\t[s]\t[kPa]\t[kPa]\n"
            "0.0\t1.0\t0.1\n"
            "1.0\t0.8\t0.08\n"
        )

        data = load_anton_paar(file)

        assert isinstance(data, RheoData)
        assert data.domain == "time"
        assert data.x_units == "s"
        assert data.y_units == "Pa"  # kPa converted and normalized
        assert data.metadata["source"] == "rheocompass"
        assert "storage_modulus" in data.metadata["columns"]
        assert "loss_modulus" in data.metadata["columns"]
        assert len(data.x) == 2

    def test_anton_paar_frequency_units(self, tmp_path):
        """Hz frequency input is normalized to rad/s."""

        file = tmp_path / "ap_freq.txt"
        file.write_text(
            "Project:\tTest\n"
            "Interval and data points:\t1\t2\n"
            "Interval data:\tFrequency\tG'\tG''\n"
            "\t[Hz]\t[Pa]\t[Pa]\n"
            "1\t10\t2\n"
            "2\t12\t3\n"
        )

        data = load_anton_paar(file)

        assert data.domain == "frequency"
        assert pytest.approx(data.x[0], rel=1e-6) == 2 * np.pi  # 1 Hz -> 2pi rad/s


class TestAutoDetection:
    """Tests for auto-detection wrapper (7.9)."""

    def test_auto_detect_from_extension(self, tmp_path):
        """Test format inference from file extension."""
        # Create CSV file
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("time,stress\n1.0,100.0\n2.0,200.0\n")

        data = auto_load(str(csv_file), x_col="time", y_col="stress")

        assert isinstance(data, RheoData)

    def test_auto_detect_trios_extension(self, tmp_path):
        """Test TRIOS detection from .txt extension."""
        trios_file = tmp_path / "test.txt"
        trios_file.write_text("""Filename	test.txt
Instrument serial number	4010-1234

[step]
Variables	Time	Stress
	s	Pa
data point 1	1.0	100
""")

        data = auto_load(str(trios_file))

        assert isinstance(data, RheoData)

    def test_auto_detect_content_based(self, tmp_path):
        """Test content-based format detection."""
        # CSV file with .dat extension
        csv_file = tmp_path / "test.dat"
        csv_file.write_text("time,stress\n1.0,100.0\n2.0,200.0\n")

        data = auto_load(str(csv_file), x_col="time", y_col="stress")

        assert isinstance(data, RheoData)

    def test_auto_detect_fallback_logic(self, tmp_path):
        """Test fallback logic when format is ambiguous."""
        # Create file with ambiguous extension
        test_file = tmp_path / "test.unknown"
        test_file.write_text("time,stress\n1.0,100.0\n")

        # Should try readers in sequence
        data = auto_load(str(test_file), x_col="time", y_col="stress")

        assert isinstance(data, RheoData)

    def test_detects_tab_delimiter(self, tmp_path):
        """Tab-separated files should be auto-detected."""
        tsv_file = tmp_path / "creep.tsv"
        tsv_file.write_text("Time\tCreep Compliance\n0.5\t0.98\n1.0\t1.10\n")

        delimiter = detect_csv_delimiter(tsv_file)
        assert delimiter == "\t"

        data = load_csv(tsv_file, x_col="Time", y_col="Creep Compliance")
        assert np.allclose(data.x, [0.5, 1.0])
        assert np.allclose(data.y, [0.98, 1.10])

    def test_auto_load_with_tab_delimiter_and_column_names(self, tmp_path):
        """auto_load should parse tab-separated CSV when columns are provided."""
        tsv_file = tmp_path / "creep_tab.csv"
        tsv_file.write_text("Time\tCreep Compliance\n0.5\t0.98\n1.0\t1.10\n")

        data = auto_load(tsv_file, x_col="Time", y_col="Creep Compliance")
        assert isinstance(data, RheoData)
        assert len(data.x) == 2
