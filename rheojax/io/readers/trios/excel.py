"""TA Instruments TRIOS Excel file reader.

This module provides a reader for TRIOS Excel exports (.xlsx, .xls) with support for:
- Metadata rows at the top of each sheet
- Multi-sheet support (one sheet per temperature/condition)
- Units in headers or separate row
- Step/Segment columns for multi-step experiments
- Complex modulus construction (G' + iG'')
- Memory-efficient read-only mode for large files

Usage:
    >>> from rheojax.io.readers.trios import load_trios_excel
    >>> data = load_trios_excel('creep_recovery.xlsx')
    >>> print(data.test_mode)  # 'creep'
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from rheojax.logging import get_logger, log_io

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    load_workbook = None

try:
    import xlrd

    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from rheojax.core.data import RheoData
from rheojax.io.readers.trios.common import (
    DataSegment,
    TRIOSFile,
    TRIOSTable,
    construct_complex_modulus,
    convert_unit,
    detect_step_column,
    detect_test_type,
    segment_to_rheodata,
    select_xy_columns,
    split_by_step,
)

logger = get_logger(__name__)

# File size threshold for read-only mode (5 MB)
LARGE_FILE_THRESHOLD_MB = 5.0

# Known unit substrings for positive-evidence unit-row detection in Excel files.
# A non-numeric row is only treated as a unit row when at least one cell contains
# one of these substrings (prevents annotation/label rows from being consumed).
_UNIT_SUBSTRINGS_EXCEL = {
    "Pa",
    "Hz",
    "rad",
    "°C",
    "°F",
    "K",
    "/s",
    "%",
    "1/",
    "mN",
    "mPa",
    "kPa",
    "MPa",
    "N·m",
    "N.m",
    "J/",
    "W/",
    "m²",
    "m2",
    "mm",
    "μm",
    "nm",
}


def _check_excel_dependencies(filepath: Path) -> None:
    """Check that required Excel libraries are available.

    Args:
        filepath: Path to Excel file (for extension check)

    Raises:
        ImportError: If required library is not installed
    """
    suffix = filepath.suffix.lower()

    if suffix == ".xlsx" and not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required to read .xlsx files. "
            "Install it with: pip install openpyxl"
        )

    if suffix == ".xls" and not HAS_XLRD:
        raise ImportError(
            "xlrd is required to read .xls files. " "Install it with: pip install xlrd"
        )


def parse_excel_metadata(
    sheet: Any,
    max_rows: int = 20,
) -> tuple[dict[str, Any], int]:
    """Extract metadata from early rows of Excel sheet.

    TRIOS Excel exports have metadata key-value pairs at the top,
    followed by the column headers.

    Args:
        sheet: openpyxl worksheet object
        max_rows: Maximum rows to scan for metadata

    Returns:
        Tuple of (metadata dict, header row index 1-based)
    """
    metadata: dict[str, Any] = {}

    # Known metadata patterns
    metadata_patterns = {
        "filename": r"^Filename",
        "instrument_serial_number": r"^Instrument serial number",
        "instrument_name": r"^Instrument name",
        "operator": r"^[Oo]perator",
        "run_date": r"^[Rr]undate",
        "sample_name": r"^Sample name",
        "geometry": r"^Geometry name",
        "geometry_type": r"^Geometry type",
        "gap": r"^Gap",
        "temperature": r"^Temperature",
        "number_of_points": r"^Number of points",
    }

    header_row = 1

    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        cell_val = sheet.cell(row=row_idx, column=1).value

        if cell_val is None:
            continue

        cell_str = str(cell_val).strip()

        # Check if this is a metadata line
        is_metadata = False
        for key, pattern in metadata_patterns.items():
            if re.match(pattern, cell_str, re.IGNORECASE):
                val_cell = sheet.cell(row=row_idx, column=2).value
                if val_cell is not None:
                    metadata[key] = str(val_cell).strip()
                is_metadata = True
                break

        if not is_metadata:
            # Check if this looks like a header row
            # Header rows have multiple non-empty text cells
            non_empty_text_cols = 0
            for col_idx in range(1, min(sheet.max_column + 1, 10)):
                cell = sheet.cell(row=row_idx, column=col_idx).value
                if cell is not None:
                    cell_val_str = str(cell).strip()
                    if cell_val_str and not _is_numeric(cell_val_str):
                        non_empty_text_cols += 1

            if non_empty_text_cols >= 3:
                header_row = row_idx
                break

    return metadata, header_row


def detect_excel_header_row(
    sheet: Any,
    start_row: int = 1,
    max_rows: int = 30,
) -> int:
    """Find data table header row in Excel sheet.

    Args:
        sheet: openpyxl worksheet object
        start_row: Row to start searching from (1-based)
        max_rows: Maximum rows to search

    Returns:
        Header row index (1-based)
    """
    for row_idx in range(start_row, min(max_rows + 1, sheet.max_row + 1)):
        cell_val = sheet.cell(row=row_idx, column=1).value

        if cell_val is None:
            continue

        cell_str = str(cell_val).strip()

        # Check for common header indicators
        if cell_str.lower() == "variables":
            return row_idx

        if cell_str.lower() == "number of points":
            # Header is typically next row after "Number of points"
            return row_idx + 1

        # Check if row has multiple text columns (header row pattern)
        text_cols = 0
        for col_idx in range(1, min(sheet.max_column + 1, 10)):
            cell = sheet.cell(row=row_idx, column=col_idx).value
            if cell is not None:
                val_str = str(cell).strip()
                if val_str and not _is_numeric(val_str):
                    text_cols += 1

        if text_cols >= 3:
            return row_idx

    return start_row


def _is_numeric(s: str) -> bool:
    """Check if string represents a numeric value."""
    try:
        float(s.replace(",", "."))
        return True
    except ValueError:
        return False


def extract_units_from_excel(
    sheet: Any,
    header_row: int,
) -> tuple[list[str], dict[str, str], bool]:
    """Parse column headers and units from Excel sheet.

    Units may be in parentheses in the header or in a separate row.

    Args:
        sheet: openpyxl worksheet object
        header_row: Header row index (1-based)

    Returns:
        Tuple of (column names list, units dict, had_unit_row)
    """
    headers: list[str] = []
    units: dict[str, str] = {}
    # R8-IO-002: track whether a dedicated unit row was found
    had_unit_row = False

    # Read header row
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx).value
        if cell is None:
            break
        header = str(cell).strip()
        headers.append(header)

        # Check for units in parentheses
        match = re.search(r"\(([^)]+)\)$", header)
        if match:
            units[header] = match.group(1)
            # Also store under name without units
            name_without_units = re.sub(r"\s*\([^)]+\)$", "", header).strip()
            units[name_without_units] = match.group(1)

    # Check if next row is a unit row
    if header_row < sheet.max_row:
        next_row = header_row + 1
        first_cell = sheet.cell(row=next_row, column=1).value

        if first_cell is not None:
            first_val = str(first_cell).strip()

            # If first cell is not numeric and not a data label, it might be units
            if not _is_numeric(first_val) and not first_val.lower().startswith("data"):
                is_unit_row = True
                # Verify a few columns (negative evidence: no numeric values)
                for col_idx in range(2, min(len(headers) + 1, 6)):
                    cell = sheet.cell(row=next_row, column=col_idx).value
                    if cell is not None and _is_numeric(str(cell).strip()):
                        is_unit_row = False
                        break

                # Require positive evidence: at least one cell contains a known unit
                # substring. This rules out annotation/label rows that happen to be
                # non-numeric (e.g. "N/A", "undefined", "--").
                if is_unit_row:
                    all_vals = [
                        str(sheet.cell(row=next_row, column=c).value or "").strip()
                        for c in range(1, min(len(headers) + 1, 7))
                    ]
                    has_unit_evidence = any(
                        any(u in v for u in _UNIT_SUBSTRINGS_EXCEL)
                        for v in all_vals
                        if v
                    )
                    if not has_unit_evidence:
                        is_unit_row = False

                if is_unit_row:
                    for col_idx, header in enumerate(headers, start=1):
                        unit_cell = sheet.cell(row=next_row, column=col_idx).value
                        if unit_cell is not None:
                            unit = str(unit_cell).strip()
                            if unit:
                                units[header] = unit
                    had_unit_row = True

    return headers, units, had_unit_row


def parse_excel_sheet(
    sheet: Any,
    filepath: str,
    table_index: int = 0,
    sheet_name: str = "Sheet1",
) -> tuple[TRIOSTable, dict[str, Any]]:
    """Parse a single Excel sheet into a TRIOSTable.

    Args:
        sheet: openpyxl worksheet object
        filepath: Original file path (for error messages)
        table_index: Index of this table
        sheet_name: Name of the sheet being parsed

    Returns:
        Tuple of (TRIOSTable, metadata dict)
    """
    # Extract metadata and find header row
    metadata, header_row = parse_excel_metadata(sheet)

    # Refine header row detection
    header_row = detect_excel_header_row(sheet, header_row)

    # Get headers and units
    headers, units, had_unit_row = extract_units_from_excel(sheet, header_row)

    if not headers:
        raise ValueError(f"No column headers found in sheet '{sheet.title}'")

    # Determine data start row
    data_start = header_row + 1

    if had_unit_row:
        # R8-IO-002: unit row already detected by extract_units_from_excel, skip it
        data_start += 1
    else:
        # Secondary check: if the first row after header is non-numeric it may be
        # a unit row that wasn't caught above (e.g. single-column files).
        # Require positive evidence to avoid consuming annotation rows.
        if data_start <= sheet.max_row:
            first_data_cell = sheet.cell(row=data_start, column=1).value
            if first_data_cell is not None:
                val_str = str(first_data_cell).strip()
                if not _is_numeric(val_str) and not val_str.lower().startswith("data"):
                    # Positive-evidence guard: at least one cell must contain a
                    # known unit substring before advancing data_start.
                    sec_vals = [
                        str(sheet.cell(row=data_start, column=c).value or "").strip()
                        for c in range(1, min(len(headers) + 1, 7))
                    ]
                    has_unit_evidence = any(
                        any(u in v for u in _UNIT_SUBSTRINGS_EXCEL)
                        for v in sec_vals
                        if v
                    )
                    if has_unit_evidence:
                        data_start += 1

    # Read data rows
    data_rows: list[list[float]] = []

    for row_idx in range(data_start, sheet.max_row + 1):
        row_data: list[float] = []
        all_empty = True

        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx).value

            if cell is None:
                row_data.append(np.nan)
            elif isinstance(cell, (int, float)):
                row_data.append(float(cell))
                all_empty = False
            else:
                val_str = str(cell).strip()
                if val_str.lower().startswith("data"):
                    # VIS-EXL-001: Use elif chain so "data"-prefixed cells
                    # append exactly one NaN and do NOT fall through to the
                    # else branch (which would append a second NaN, making
                    # row_data longer than headers and causing the row to be
                    # silently discarded at the length-mismatch guard below).
                    row_data.append(np.nan)
                elif not val_str:
                    row_data.append(np.nan)
                else:
                    try:
                        row_data.append(float(val_str.replace(",", ".")))
                        all_empty = False
                    except ValueError:
                        row_data.append(np.nan)

        # Skip empty rows
        if all_empty:
            continue

        if row_data:
            # Ensure row matches header length
            if len(row_data) == len(headers):
                data_rows.append(row_data)
            elif len(row_data) < len(headers):
                # Pad with NaN
                row_data.extend([np.nan] * (len(headers) - len(row_data)))
                data_rows.append(row_data)

    if not data_rows:
        raise ValueError(f"No data rows found in sheet '{sheet.title}'")

    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=headers)

    # Remove columns that are all NaN
    n_cols_before = len(df.columns)
    df = df.dropna(axis=1, how="all")
    n_dropped = n_cols_before - len(df.columns)
    if n_dropped > 0:
        logger.info(
            "Dropped all-NaN columns",
            n_dropped=n_dropped,
            remaining=len(df.columns),
        )

    # Detect step column
    step_col = detect_step_column(df)
    step_values = None
    if step_col:
        n_step_total = len(df[step_col])
        step_values = df[step_col].dropna().unique().astype(int).tolist()
        n_step_nan = n_step_total - len(df[step_col].dropna())
        if n_step_nan > 0:
            logger.debug(
                "Step values with NaN excluded",
                n_nan=n_step_nan,
                n_total=n_step_total,
            )

    table = TRIOSTable(
        table_index=table_index,
        header=list(df.columns),
        units={k: v for k, v in units.items() if k in df.columns},
        df=df,
        step_values=step_values,
        sheet_name=sheet_name,
    )

    return table, metadata


def parse_trios_excel(
    filepath: str | Path,
    *,
    sheet_name: str | int | None = None,
    read_only: bool | None = None,
) -> TRIOSFile:
    """Low-level Excel parser returning raw TRIOSFile structure.

    For advanced users who need access to raw tables and metadata
    before RheoData conversion.

    Args:
        filepath: Path to TRIOS Excel file
        sheet_name: Sheet to parse (None=first, "all"=all, int=index, str=name)
        read_only: Use read-only mode for large files (None=auto)

    Returns:
        TRIOSFile with parsed tables and metadata

    Raises:
        FileNotFoundError: File does not exist
        ValueError: Sheet not found or no data
    """
    filepath = Path(filepath)

    if not filepath.exists():
        logger.error("File not found", filepath=str(filepath))
        raise FileNotFoundError(f"File not found: {filepath}")

    _check_excel_dependencies(filepath)

    # Determine if we should use read-only mode
    file_size_mb = filepath.stat().st_size / (1024 * 1024)
    if read_only is None:
        read_only = file_size_mb > LARGE_FILE_THRESHOLD_MB
        if read_only:
            logger.debug(
                "Using read-only mode for large file",
                filepath=str(filepath),
                file_size_mb=round(file_size_mb, 2),
            )

    # Load workbook
    suffix = filepath.suffix.lower()
    logger.debug("Loading Excel workbook", filepath=str(filepath), format=suffix)

    with log_io(logger, "read", filepath=str(filepath)) as io_ctx:
        io_ctx["format"] = suffix
        io_ctx["file_size_mb"] = round(file_size_mb, 2)

        wb = None
        if suffix == ".xlsx":
            wb = load_workbook(filepath, read_only=read_only, data_only=True)
            sheet_names = wb.sheetnames
        elif suffix == ".xls":
            # xlrd handles .xls files
            wb = xlrd.open_workbook(str(filepath))
            sheet_names = wb.sheet_names()
        else:
            logger.error(
                "Unsupported Excel format", filepath=str(filepath), format=suffix
            )
            raise ValueError(f"Unsupported Excel format: {suffix}")

        try:
            logger.debug(
                "Workbook loaded",
                filepath=str(filepath),
                sheet_count=len(sheet_names),
                sheets=sheet_names,
            )

            # Determine which sheets to parse
            if sheet_name == "all":
                sheets_to_parse = list(range(len(sheet_names)))
            elif sheet_name is None:
                sheets_to_parse = [0]
            elif isinstance(sheet_name, int):
                if sheet_name < 0 or sheet_name >= len(sheet_names):
                    logger.error(
                        "Sheet index out of range",
                        filepath=str(filepath),
                        sheet_index=sheet_name,
                        available=f"0-{len(sheet_names) - 1}",
                    )
                    raise ValueError(
                        f"Sheet index {sheet_name} out of range. "
                        f"Available: 0-{len(sheet_names) - 1}"
                    )
                sheets_to_parse = [sheet_name]
            elif isinstance(sheet_name, str):
                if sheet_name not in sheet_names:
                    logger.error(
                        "Sheet not found",
                        filepath=str(filepath),
                        sheet_name=sheet_name,
                        available=sheet_names,
                    )
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. " f"Available: {sheet_names}"
                    )
                sheets_to_parse = [sheet_names.index(sheet_name)]
            else:
                logger.error(  # type: ignore[unreachable]
                    "Invalid sheet_name type",
                    filepath=str(filepath),
                    sheet_name_type=type(sheet_name).__name__,
                )
                raise ValueError(f"Invalid sheet_name type: {type(sheet_name)}")

            # Parse sheets
            tables: list[TRIOSTable] = []
            global_metadata: dict[str, Any] = {}

            for idx, sheet_idx in enumerate(sheets_to_parse):
                sheet_name_str = sheet_names[sheet_idx]
                logger.debug(
                    "Parsing sheet",
                    filepath=str(filepath),
                    sheet_name=sheet_name_str,
                    sheet_index=sheet_idx,
                )

                if suffix == ".xlsx":
                    sheet = wb[sheet_name_str]
                else:
                    sheet = _XlrdSheetWrapper(wb.sheet_by_index(sheet_idx))

                try:
                    table, sheet_metadata = parse_excel_sheet(
                        sheet, str(filepath), idx, sheet_name=sheet_name_str
                    )
                except Exception:
                    logger.error(
                        "Failed to parse sheet",
                        filepath=str(filepath),
                        sheet_name=sheet_name_str,
                        exc_info=True,
                    )
                    raise

                tables.append(table)
                logger.debug(
                    "Sheet parsed",
                    filepath=str(filepath),
                    sheet_name=sheet_name_str,
                    rows=len(table.df),
                    columns=len(table.df.columns),
                )

                # Merge metadata (first sheet metadata is primary)
                if idx == 0:
                    global_metadata = sheet_metadata
                else:
                    # Store per-sheet metadata
                    global_metadata[f"sheet_{sheet_idx}_metadata"] = sheet_metadata
        finally:
            if wb is not None:
                if suffix == ".xlsx":
                    wb.close()
                elif suffix == ".xls" and hasattr(wb, "release_resources"):
                    wb.release_resources()

        if not tables:
            logger.error("No data tables found", filepath=str(filepath))
            raise ValueError(f"No data tables found in {filepath}")

        io_ctx["sheets_parsed"] = len(tables)
        io_ctx["total_rows"] = sum(len(t.df) for t in tables)

    return TRIOSFile(
        filepath=str(filepath),
        format="excel",
        metadata=global_metadata,
        tables=tables,
        encoding="utf-8",  # Excel handles encoding internally
        decimal_separator=".",
    )


class _XlrdSheetWrapper:
    """Wrapper to provide openpyxl-like interface for xlrd sheets."""

    def __init__(self, sheet: Any) -> None:
        self._sheet = sheet
        self.title = sheet.name
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell(self, row: int, column: int) -> Any:
        """Get cell value (1-based indexing like openpyxl)."""

        class CellWrapper:
            def __init__(self, value: Any) -> None:
                self.value = value

        try:
            r, c = row - 1, column - 1
            cell_type = self._sheet.cell_type(r, c)
            if cell_type in (0, 5):  # XL_CELL_EMPTY or XL_CELL_ERROR
                return CellWrapper(None)
            return CellWrapper(self._sheet.cell_value(r, c))
        except IndexError:
            return CellWrapper(None)


def load_trios_excel(
    filepath: str | Path,
    *,
    sheet_name: str | int | None = None,
    return_all_segments: bool = False,
    test_mode: str | None = None,
    validate: bool = True,
) -> RheoData | list[RheoData]:
    """Load TRIOS Excel export file (.xlsx, .xls).

    Handles TRIOS-specific Excel format with:
    - Metadata in early rows
    - Data table with headers
    - Multi-sheet support (one sheet per temperature)

    Args:
        filepath: Path to TRIOS Excel file
        sheet_name: Sheet to load (None=first, "all"=all sheets, int=index, str=name)
        return_all_segments: Return list for multi-step files
        test_mode: Override auto-detection ("creep", "relaxation", "oscillation", "rotation")
        validate: Validate RheoData on creation

    Returns:
        Single RheoData for single-sheet files
        List of RheoData when sheet_name="all" or multiple sheets/segments

    Raises:
        FileNotFoundError: File does not exist
        ValueError: Sheet not found or no data

    Example:
        >>> data = load_trios_excel('creep_recovery.xlsx')
        >>> print(data.test_mode)  # 'creep'

        >>> all_sheets = load_trios_excel('multi_temp.xlsx', sheet_name='all')
        >>> for d in all_sheets:
        ...     print(d.metadata.get('temperature'))
    """
    # Parse Excel file
    trios_file = parse_trios_excel(filepath, sheet_name=sheet_name)

    # Convert tables to RheoData
    rheo_data_list: list[RheoData] = []

    for table in trios_file.tables:
        df = table.df
        units = table.units
        sheet_name_str = table.sheet_name

        # Detect or use provided test mode
        detected_mode = test_mode or detect_test_type(df)

        # Check for step column and split if needed
        step_col = detect_step_column(df)
        segments = (
            [df]
            if not step_col or not return_all_segments
            else split_by_step(df, step_col)
        )

        for seg_idx, seg_df in enumerate(segments):
            # Select x/y columns
            x_col, y_col, y2_col = select_xy_columns(seg_df, detected_mode)

            if x_col is None or y_col is None:
                msg = (
                    f"Skipping TRIOS Excel segment {seg_idx} in sheet "
                    f"'{sheet_name_str}': could not determine x/y columns. "
                    f"Available columns: {list(seg_df.columns)}"
                )
                warnings.warn(msg, stacklevel=2)
                logger.warning(
                    f"Could not determine x/y columns for segment {seg_idx} "
                    f"in sheet '{sheet_name_str}'. "
                    f"Available columns: {list(seg_df.columns)}"
                )
                continue

            # Extract data
            try:
                x_data = seg_df[x_col].values.astype(float)
            except (ValueError, TypeError) as e:
                raise ValueError(
                    f"Column '{x_col}' contains non-numeric data that cannot be converted to float. "
                    f"Sample values: {seg_df[x_col].head(3).tolist()}"
                ) from e

            # Get units
            x_units = units.get(x_col, "")
            y_units = units.get(y_col, "Pa")

            # Handle complex modulus case
            if y2_col is not None:
                try:
                    y_real = seg_df[y_col].values.astype(float)
                except (ValueError, TypeError) as e:
                    raise ValueError(
                        f"Column '{y_col}' contains non-numeric data that cannot be converted to float. "
                        f"Sample values: {seg_df[y_col].head(3).tolist()}"
                    ) from e
                try:
                    y_imag = seg_df[y2_col].values.astype(float)
                except (ValueError, TypeError) as e:
                    raise ValueError(
                        f"Column '{y2_col}' contains non-numeric data that cannot be converted to float. "
                        f"Sample values: {seg_df[y2_col].head(3).tolist()}"
                    ) from e

                # Convert units if needed
                y_units_orig = units.get(y_col, "Pa")
                y2_units_orig = units.get(y2_col, "Pa")
                y_real, _ = convert_unit(y_real, y_units_orig, "Pa")
                y_imag, _ = convert_unit(y_imag, y2_units_orig, "Pa")

                # Construct complex modulus
                y_data = construct_complex_modulus(y_real, y_imag)
                y_units = "Pa"
                is_complex = True
            else:
                try:
                    y_data = seg_df[y_col].values.astype(float)
                except (ValueError, TypeError) as e:
                    raise ValueError(
                        f"Column '{y_col}' contains non-numeric data that cannot be converted to float. "
                        f"Sample values: {seg_df[y_col].head(3).tolist()}"
                    ) from e
                is_complex = False

            # Convert x units (e.g., Hz to rad/s)
            if detected_mode == "oscillation":
                x_data, x_units = convert_unit(x_data, x_units, "rad/s")

            # Remove NaN values
            if is_complex:
                valid_mask = ~(
                    np.isnan(x_data)
                    | np.isnan(np.real(y_data))
                    | np.isnan(np.imag(y_data))
                )
            else:
                valid_mask = ~(np.isnan(x_data) | np.isnan(y_data))

            x_data = x_data[valid_mask]
            y_data = y_data[valid_mask]

            if len(x_data) == 0:
                continue

            # Determine default x_units based on test mode
            if not x_units:
                if detected_mode == "oscillation":
                    x_units = "rad/s"
                elif detected_mode == "rotation":
                    x_units = "1/s"
                else:
                    x_units = "s"

            # Build metadata
            seg_metadata = trios_file.metadata.copy()
            seg_metadata["test_mode"] = detected_mode
            seg_metadata["source_format"] = "excel"
            seg_metadata["sheet_name"] = sheet_name_str
            seg_metadata["x_column"] = x_col
            seg_metadata["y_column"] = y_col
            if y2_col:
                seg_metadata["y2_column"] = y2_col
            seg_metadata["is_complex"] = is_complex

            # Create DataSegment and convert to RheoData
            segment = DataSegment(
                segment_index=seg_idx,
                test_mode=detected_mode,
                x_data=x_data,
                y_data=y_data,
                x_column=x_col,
                y_column=y_col,
                x_units=x_units,
                y_units=y_units,
                is_complex=is_complex,
                metadata=seg_metadata,
            )

            rheo_data = segment_to_rheodata(segment, validate=validate)
            rheo_data_list.append(rheo_data)

    if not rheo_data_list:
        raise ValueError(f"No valid data segments could be parsed from {filepath}")

    # Return single or list
    if len(rheo_data_list) == 1 and not return_all_segments and sheet_name != "all":
        return rheo_data_list[0]
    return rheo_data_list
